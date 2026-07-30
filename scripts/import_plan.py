#!/usr/bin/env python3
"""
يستورد خطة المحتوى الموحّدة من الإكسل إلى مخطط النظام.

    python3 scripts/import_plan.py "path/to/زيادة-خطة-المحتوى-الموحدة.xlsx"

لماذا الاستيراد لا إعادة الكتابة:
  ملف الإكسل أوسع من خطتي في أمور جوهرية: أفق ستة عشر شهراً بدل
  ثلاثة، وواحد وخمسون عنصراً، وأربع منصات فيها تيك توك وواتساب،
  ومناسبات محلية غابت عني أهمها «فترة الرواتب» شهرياً. فهو مصدر
  الحقيقة للمحتوى، والنظام محرّك تنفيذ لا مصدر أفكار.

ما يفعله:
  1. يقرأ الخط الزمني وصفحات المنصات
  2. يحوّل تواريخ العربية إلى ISO ويثبّتها على أحد
  3. يستخرج الصيغة لكل منصة (عدد شرائح الكاروسيل، طول الريلز، عدد
     تغريدات الثريد) فالصيغة تحدد القالب المطلوب
  4. يكتب content/imported.yml ويطبع تقريراً بما يحتاج يداً بشرية

ما لا يفعله:
  لا يخترع نص شريحة ولا كابشن. ما لم يوجد في الإكسل يبقى فارغاً
  ومعلّماً في التقرير، لأن ملء فراغ بمحتوى مولَّد يفقد الملف قيمته
  كمصدر تعلّم.
"""
import re
import sys
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parent.parent

MONTHS = {
    "يناير": 1, "فبراير": 2, "مارس": 3, "أبريل": 4, "ابريل": 4, "مايو": 5,
    "يونيو": 6, "يوليو": 7, "أغسطس": 8, "اغسطس": 8, "سبتمبر": 9,
    "أكتوبر": 10, "اكتوبر": 10, "نوفمبر": 11, "ديسمبر": 12,
}

PRIORITY = {"1 حرجة": "critical", "2 عالية": "high", "3 متوسطة": "medium"}

# تحويل مدة التحضير إلى أيام. الحد الأدنى ثلاثة أيام في الملف الأصلي،
# والأقصى ستة أسابيع للجمعة البيضاء، وهذا فرق حقيقي في التخطيط.
PREP = {
    "3 أيام": 3, "أسبوع": 7, "أسبوعان": 14, "3 أسابيع": 21,
    "شهر": 30, "6 أسابيع": 42,
}


def parse_ar_date(text):
    """يحوّل '2 أغسطس 2026' إلى date. يُرجع None لما ليس تاريخاً."""
    if not text:
        return None
    m = re.search(r"(\d{1,2})\s+(\S+)\s+(\d{4})", str(text))
    if not m:
        return None
    day, mon, year = m.group(1), m.group(2), m.group(3)
    if mon not in MONTHS:
        return None
    return date(int(year), MONTHS[mon], int(day))


def to_sunday(d: date) -> date:
    """
    يثبّت التاريخ على أحد. جدول النشر مبني على أسبوع العمل السعودي
    الذي يبدأ الأحد، والملف الأصلي أغلب تواريخه أحد أصلاً. الإزاحة
    تُسجَّل في التقرير حتى لا تمر صامتة.
    """
    return d - timedelta(days=(d.weekday() - 6) % 7)


def parse_format(text):
    """
    يستخرج الصيغة والحجم. الصيغة تحدد القالب المطلوب:
    كاروسيل 8 شرائح يحتاج كانفس ثمانية، والريلز يحتاج فيديو لا صورة.
    """
    if not text or str(text).strip() in ("—", "-", ""):
        return None
    t = str(text).strip()

    m = re.search(r"كاروسيل\s*(\d+)", t)
    if m:
        return {"kind": "carousel", "slides": int(m.group(1)), "raw": t}
    m = re.search(r"ثريد\s*(\d+)", t)
    if m:
        return {"kind": "thread", "tweets": int(m.group(1)), "raw": t}
    m = re.search(r"(\d+)\s*ثانية", t)
    if m:
        kind = "reels" if "ريلز" in t else "video"
        return {"kind": kind, "seconds": int(m.group(1)), "raw": t}
    if "استطلاع" in t:
        return {"kind": "poll", "raw": t}
    if "رسالة" in t:
        return {"kind": "message", "raw": t}
    return {"kind": "post", "raw": t}


def slug_from(idea: str, used: set) -> str:
    """معرّف لاتيني مستقر. العربية في المعرّف تكسر أسماء الملفات."""
    base = re.sub(r"[^\w\u0600-\u06ff]+", "-", str(idea).strip())[:34].strip("-")
    base = re.sub(r"[\u0600-\u06ff]", "", base).strip("-") or "unit"
    n, out = 1, base
    while out in used:
        n += 1
        out = "%s-%d" % (base, n)
    used.add(out)
    return out


def read_platform_copy(wb) -> dict:
    """
    يجمع النصوص الجاهزة من صفحات المنصات، مفهرسةً بالرقم التسلسلي
    فهو الرابط الوحيد الموثوق بين الخط الزمني وصفحات المنصات.
    """
    out = {}
    sheets = {
        "إنستقرام": ("instagram", {"cover": 6, "slides": 7, "caption": 8, "cta": 9}),
        "إكس": ("twitter", {}),
        "تيك توك": ("tiktok", {}),
        "واتساب": ("whatsapp", {"message": 6, "timing": 7}),
    }
    for name, (key, cols) in sheets.items():
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for r in ws.iter_rows(min_row=4, values_only=True):
            if not r or r[0] is None:
                continue
            try:
                idx = int(float(r[0]))
            except (TypeError, ValueError):
                continue
            rec = out.setdefault(idx, {})
            body = {}
            for field, ci in cols.items():
                if ci < len(r) and r[ci] not in (None, "", "—"):
                    body[field] = str(r[ci]).strip()
            if not cols:
                # صفحة بلا خريطة أعمدة: نأخذ أطول خلية نصية كنص جاهز
                texts = [str(c).strip() for c in r[5:] if isinstance(c, str) and len(str(c)) > 40]
                if texts:
                    body["text"] = max(texts, key=len)
            if body:
                rec[key] = body
    return out


def main():
    if len(sys.argv) < 2:
        raise SystemExit("مرر مسار ملف الإكسل.")
    wb = openpyxl.load_workbook(sys.argv[1], data_only=True)

    if "الخط الزمني" not in wb.sheetnames:
        raise SystemExit("لا توجد ورقة 'الخط الزمني'. تحقق من الملف.")

    copy = read_platform_copy(wb)
    ws = wb["الخط الزمني"]

    units, notes, used = [], [], set()
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or r[1] is None:
            continue
        d = parse_ar_date(r[1])
        if not d:
            notes.append("سطر بتاريخ غير مقروء: %s" % str(r[1])[:60])
            continue

        # بعض الصفوف رأس قسم لا عنصر: خانة الرقم فيها اسم مناسبة
        try:
            idx = int(float(r[0]))
        except (TypeError, ValueError):
            notes.append('صف بلا رقم تسلسلي، تُخطّي: %s' % str(r[0])[:50])
            continue
        sunday = to_sunday(d)
        if sunday != d:
            notes.append("%s أُزيح إلى أحد %s" % (d.isoformat(), sunday.isoformat()))

        idea = str(r[5] or "").strip()
        occ = str(r[4] or "").strip()
        is_occ = str(r[3] or "").strip() == "مناسبة"

        formats = {}
        for ci, key in ((6, "twitter"), (7, "instagram"), (8, "tiktok"), (9, "whatsapp")):
            f = parse_format(r[ci] if ci < len(r) else None)
            if f:
                formats[key] = f

        unit = {
            "date": sunday.isoformat(),
            "slug": slug_from(idea or occ, used),
            "source_row": idx,
            "act": "occasion" if is_occ else "capability",
            "idea": idea,
            "priority": PRIORITY.get(str(r[10] or "").strip(), "high"),
            "prep_days": PREP.get(str(r[11] or "").strip(), 3),
            "formats": formats,
        }
        if is_occ:
            unit["occasion"] = re.sub(r"[\U0001f000-\U0001ffff\u2600-\u27bf]", "", occ).strip()
            if "متكرر شهري" in occ:
                unit["recurring"] = "monthly"

        if idx in copy:
            unit["source_copy"] = copy[idx]
        else:
            notes.append("الصف %s بلا نص جاهز في أي صفحة منصة" % idx)

        units.append(unit)

    # تواريخ مكررة: الملف ينشر ثلاث أفكار في نفس الأحد، والنظام
    # يقبل وحدة واحدة لكل تاريخ. هذا أهم تعارض بنيوي بين الاثنين.
    from collections import Counter
    dupes = {k: v for k, v in Counter(u["date"] for u in units).items() if v > 1}

    out = {
        "source": Path(sys.argv[1]).name,
        "imported_units": len(units),
        "units": units,
    }
    path = ROOT / "content" / "imported.yml"
    path.write_text(yaml.dump(out, allow_unicode=True, sort_keys=False, width=100),
                    encoding="utf-8")

    print("استُوردت %d وحدة إلى %s" % (len(units), path.relative_to(ROOT)))
    print()
    print("المدى: %s إلى %s" % (units[0]["date"], units[-1]["date"]))
    print("المناسبات: %d | الدائم: %d"
          % (sum(1 for u in units if u["act"] == "occasion"),
             sum(1 for u in units if u["act"] != "occasion")))
    print("الأولويات: %s" % dict(Counter(u["priority"] for u in units)))
    print("الصيغ: %s" % dict(Counter(
        f["kind"] for u in units for f in u["formats"].values())))
    print()

    if dupes:
        print("تعارض بنيوي: %d تاريخ فيه أكثر من وحدة." % len(dupes))
        print("الملف ينشر ثلاث أفكار في الأحد الواحد، والنظام يقبل واحدة.")
        print("القرار مطلوب منك: نوسّع النظام لعدة وحدات في اليوم،")
        print("أو نوزّع الأفكار على أحد وثلاثاء وخميس.")
        for k in sorted(dupes)[:5]:
            print("   %s ← %d وحدات" % (k, dupes[k]))
        print()

    if notes:
        print("يحتاج انتباهاً (%d):" % len(notes))
        for n in notes[:12]:
            print("   - " + n)
        if len(notes) > 12:
            print("   ... و%d أخرى" % (len(notes) - 12))


if __name__ == "__main__":
    main()
